from openpyxl import Workbook, load_workbook

class UseExcel():
    def __init__(self,file_name=None,sheet_id=None):
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id	
        else:
            self.file_name = './static/case1.xlsx'
            self.sheet_id = 0
            self.data = self.get_data()
    def get_data(self):
        wb = load_workbook('./static/case1.xlsx',keep_vba=True)
        ws = wb.active
        return ws
	#获取单元格的行数
    def get_lines(self):
        return self.data.max_row

    def get_content(self):
        content = []
        for row in self.data.rows:
            line = [col.value for col in row]
            content.append(line)
        return content
	#获取某一个单元格的内容
    def get_cell_value(self,row=1,col=1):
        return self.data.cell(row,col).value
        

if __name__ == '__main__':
    client = UseExcel()
    print(client.get_content())